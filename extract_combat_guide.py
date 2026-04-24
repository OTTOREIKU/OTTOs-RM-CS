import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import json
import re
import openpyxl

XLSX = 'RMU - Character Sheet 3.0 - _blank_.xlsx'
OUTPUT = 'src/data/combat_guide.json'

wb = openpyxl.load_workbook(XLSX, data_only=True)

# ─────────────────────────────────────────────
# SHEET: Combat Action Reference
# ─────────────────────────────────────────────

def parse_action(text):
    """Parse an action string like 'Normal Attack (4AP)' -> {action, ap, modifier, note}"""
    if not text or not isinstance(text, str):
        return None
    text = text.strip()
    # Extract AP
    ap = None
    modifier = None
    note = None
    ap_match = re.search(r'\((\d+)AP[^)]*\)', text)
    if ap_match:
        ap = int(ap_match.group(1))
    # Extract modifier like -25, -50, -75
    mod_match = re.search(r'([+-]\d+)\s*(?:all other actions)?', text)
    if mod_match and '-' in mod_match.group(1):
        # Only if it's part of the action name pattern (e.g. "Rushed -25")
        if re.search(r'(?:Rushed|Snap|rushed|snap)\s*[A-Za-z]*\s*(-\d+)', text):
            mod_match2 = re.search(r'(-\d+)', text)
            if mod_match2:
                modifier = int(mod_match2.group(1))
    # Movement penalty note
    if 'all other actions' in text.lower():
        penalty_match = re.search(r'-(\d+)\s+all other actions', text, re.IGNORECASE)
        if penalty_match:
            note = f'-{penalty_match.group(1)} all other actions'
    # Clean action name (remove AP portion)
    action_name = re.sub(r'\s*\([^)]*AP[^)]*\)', '', text).strip()
    action_name = re.sub(r'\s*-\d+\s*all other actions', '', action_name, flags=re.IGNORECASE).strip()
    return {"action": action_name, "ap": ap, "modifier": modifier, "note": note}


ws_car = wb['Combat Action Reference']

# Column layout (1-indexed): B=Melee, F=Ranged, J=Movement, N=Movement(no AP), R=Other movements, V=Actions, Z=Spells
col_map = {
    "Melee Attack": 2,
    "Ranged Attack": 6,
    "Movement": 10,
    "Movement (Free)": 14,
    "Other movements": 18,
    "Actions": 22,
    "Spells": 26,
}

categories = {cat: [] for cat in col_map}

# AP by status table is in row 23 (actual Excel row), cols L-Z (every 2 cols)
# Row 22: headers "Instant Action", "Conc." repeating
# Row 23: "Held 1AP", "Slowed 2AP\n...", etc. at cols 12,14,16,18,20,22,24,26
ap_by_status = []
row23 = list(ws_car.iter_rows(min_row=23, max_row=23, values_only=True))[0]

# Status cells appear at even-indexed 0-based positions: 11,13,15,17,19,21,23,25
for col_idx in [11, 13, 15, 17, 19, 21, 23, 25]:
    val = row23[col_idx]  # 0-indexed
    if val and isinstance(val, str):
        lines = val.strip().split('\n')
        status_line = lines[0].strip()
        ap_match = re.search(r'(\w[\w ]*?)\s+(\d+)AP', status_line)
        if ap_match:
            status_name = ap_match.group(1).strip()
            ap_count = int(ap_match.group(2))
            note_parts = lines[1:] if len(lines) > 1 else []
            ap_by_status.append({
                "status": status_name,
                "ap": ap_count,
                "note": '; '.join(p.strip() for p in note_parts if p.strip()) or None
            })

# Collect action rows (rows 4-22; row 2=header, row 23+ is AP status table)
for row in ws_car.iter_rows(min_row=4, max_row=22, values_only=True):
    for cat, col in col_map.items():
        cell_val = row[col - 1]
        if cell_val and isinstance(cell_val, str):
            parsed = parse_action(cell_val)
            if parsed:
                categories[cat].append(parsed)

# Also collect the "Other movements" note from row 8 (it has a note in col N area)
# Row 8 col 14 area has "Movement penalties for moving without AP..."
other_move_note = None
for row in ws_car.iter_rows(min_row=2, max_row=15, values_only=True):
    for col_idx in range(13, 17):
        cell = row[col_idx]
        if cell and isinstance(cell, str) and 'Movement penalties' in cell:
            other_move_note = cell.strip()

# Collect rules notes (long text cells)
notes = []
for row in ws_car.iter_rows(min_row=16, max_row=33, values_only=True):
    for cell in row:
        if cell and isinstance(cell, str) and len(cell) > 50:
            cleaned = cell.strip()
            if cleaned not in notes:
                notes.append(cleaned)

actions_out = [
    {"category": cat, "entries": entries}
    for cat, entries in categories.items()
]

round_data = {
    "duration_seconds": 5,
    "ap_by_status": ap_by_status,
    "notes": notes
}

# ─────────────────────────────────────────────
# SHEET: WT_Ref
# ─────────────────────────────────────────────

ws_wt = wb['WT_Ref']

# --- Crit size adjustments ---
# Size table rows 2-11 cols Q(17)-U(21)
# Original Crit adjustment matrix rows 2-14 cols W(23)-AM(39)
# Row 2 = column adjustment headers (-9 to +8)
# Row 3 = row adjustment index 1 (original crit level 1?)
# Rows 4-14 = original crit letters Z,A,B,...,J with adjusted results

# Size description table (attacker/defender adjustments)
size_rows = []
for i, row in enumerate(ws_wt.iter_rows(min_row=2, max_row=11, min_col=17, max_col=21, values_only=True), 2):
    size_desc, attack_size, hits_mult, att_crit_adj, def_crit_adj = row
    if size_desc:
        size_rows.append({
            "size": str(size_desc),
            "attack_size": int(attack_size) if attack_size is not None else None,
            "hits_multiplier": float(hits_mult) if hits_mult is not None else None,
            "attacker_crit_adjustment": int(att_crit_adj) if att_crit_adj is not None else None,
            "defender_crit_adjustment": int(def_crit_adj) if def_crit_adj is not None else None,
        })

# Crit result lookup matrix: cols W(23)-AM(39), rows 2-14
# Row 2 header row: col W = None (header label), X..AM = adjustment offsets -9..+8
# Row 3: offset row index (row index 1), col X onward = adjusted crit letters
# Rows 4-14: original crit = letter, adjusted results across columns
from openpyxl.utils import column_index_from_string
w_col = column_index_from_string('W')

crit_adj_offsets_row = list(ws_wt.iter_rows(min_row=2, max_row=2, min_col=w_col+1, max_col=w_col+18, values_only=True))[0]
crit_adj_offsets = [int(v) for v in crit_adj_offsets_row if v is not None]

crit_matrix = []
# rows 4-14 (0-indexed row 3-13 in loop, Excel rows 4-14)
for row in ws_wt.iter_rows(min_row=4, max_row=14, min_col=w_col, max_col=w_col+18, values_only=True):
    orig_crit = row[0]
    if orig_crit:
        results = list(row[1:len(crit_adj_offsets)+1])
        crit_matrix.append({
            "original_crit": str(orig_crit),
            "adjustments": {str(crit_adj_offsets[i]): str(results[i]) for i in range(len(crit_adj_offsets)) if i < len(results) and results[i] is not None}
        })

crit_size_out = {
    "note": "Attacker crit adjustment based on size. Use attacker_crit_adjustment - defender_crit_adjustment to get net offset into the matrix.",
    "size_table": size_rows,
    "crit_result_matrix": crit_matrix
}

# --- Vision penalties ---
# Rows 35-41, cols W(23)-AC(29)
# Header row 34: W=Vision Conditions, X=Vision Required (Normal), Y=Vision Helpful (Normal), Z=Vision Required (Nightvision), AA=Vision Helpful (NV), AB=Vision Required (DV), AC=Vision Helpful (DV)
vision_penalties = []
for row in ws_wt.iter_rows(min_row=35, max_row=41, min_col=23, max_col=29, values_only=True):
    condition, norm_req, norm_help, nv_req, nv_help, dv_req, dv_help = row
    if condition:
        vision_penalties.append({
            "condition": str(condition),
            "normal_vision_required": int(norm_req) if norm_req is not None else 0,
            "normal_vision_helpful": int(norm_help) if norm_help is not None else 0,
            "nightvision_required": int(nv_req) if nv_req is not None else 0,
            "nightvision_helpful": int(nv_help) if nv_help is not None else 0,
            "darkvision_required": int(dv_req) if dv_req is not None else 0,
            "darkvision_helpful": int(dv_help) if dv_help is not None else 0,
        })

# --- Slaying bonuses ---
# Rows 45-50, cols W(23)-X(24)
slaying_bonuses = []
for row in ws_wt.iter_rows(min_row=45, max_row=50, min_col=23, max_col=24, values_only=True):
    tier, bonus = row
    if tier is not None:
        slaying_bonuses.append({
            "tier": str(tier),
            "bonus": int(bonus) if bonus is not None else 0
        })

# --- Weapon categories multi-attack penalty table ---
# Rows 101-116, cols A-H
# Row 101: headers Melee, Ranged, Shield, Unarmed
# Row 102: sub-header <weapon skill>
# Rows 103-116: category name, penalty, category name, penalty, ... (pairs)
weapon_multi_attack = []
col_groups = [
    ("Melee", 1, 2),
    ("Ranged", 3, 4),
    ("Shield", 5, 6),
    ("Unarmed", 7, 8),
]
for row in ws_wt.iter_rows(min_row=103, max_row=116, values_only=True):
    for group_name, cat_col, pen_col in col_groups:
        cat = row[cat_col - 1]
        pen = row[pen_col - 1]
        if cat and isinstance(cat, str):
            weapon_multi_attack.append({
                "group": group_name,
                "skill": str(cat),
                "multi_attack_penalty": int(pen) if pen is not None else None
            })
        elif pen is not None and cat is None:
            # Penalty without category name - last few rows have None cats with same penalty
            pass  # skip placeholder rows

# ─────────────────────────────────────────────
# SHEET: MT_Ref
# ─────────────────────────────────────────────

ws_mt = wb['MT_Ref']

# --- Mount stats ---
# Rows 2-14, cols A-J (1-10): Mount, Weight, BMR, AT, Hits, Endurance, Load Bonus, OB, DB, Crit
mount_stats = []
for row in ws_mt.iter_rows(min_row=2, max_row=14, min_col=1, max_col=10, values_only=True):
    name = row[0]
    if name and isinstance(name, str):
        mount_stats.append({
            "name": str(name),
            "weight": row[1],
            "bmr": row[2],
            "at": row[3],
            "hits": row[4],
            "endurance": row[5],
            "load_bonus": row[6],
            "ob": row[7],
            "db": row[8],
            "crit_type": str(row[9]) if row[9] else None,
        })

# --- Endurance check results ---
# Rows 2-7, col 13 (M) = tier label, col 14 (N) = narrative text
endurance_results = []
for row in ws_mt.iter_rows(min_row=2, max_row=7, min_col=13, max_col=14, values_only=True):
    tier_label, narrative = row
    if tier_label and isinstance(tier_label, str):
        result_text = narrative.strip() if narrative and isinstance(narrative, str) else None
        endurance_results.append({
            "tier": str(tier_label),
            "result": result_text
        })

# --- Recovery table ---
# Rows 26-202, cols A-G
# Col A = roll value (0-176), Cols B-G = rounds to recover per injury type
# Injury types: Bone, Cuts & Burns, Muscle/Tendon, Organ, Poison/Disease, Skin/Tissue
# The data groups into tiers based on where values change:
# roll 0: special, rolls 1-50, 51-75, 76-100, 101-125, 126-175, 176+
injury_types = ["Bone", "Cuts & Burns", "Muscle/Tendon", "Organ", "Poison/Disease", "Skin/Tissue"]

# Build condensed recovery tiers
recovery_tiers = {}
for row in ws_mt.iter_rows(min_row=26, max_row=202, min_col=1, max_col=7, values_only=True):
    roll = row[0]
    if roll is None:
        continue
    roll = int(roll)
    values = tuple(int(v) if v is not None else None for v in row[1:7])
    if values not in recovery_tiers.values():
        recovery_tiers[roll] = values

# Identify the tier break points
tier_breaks = []
prev_vals = None
for row in ws_mt.iter_rows(min_row=26, max_row=202, min_col=1, max_col=7, values_only=True):
    roll = row[0]
    if roll is None:
        continue
    roll = int(roll)
    vals = tuple(row[1:7])
    if vals != prev_vals:
        tier_breaks.append((roll, vals))
        prev_vals = vals

# Build recovery output with roll ranges
recovery = []
for i, (start_roll, vals) in enumerate(tier_breaks):
    if i + 1 < len(tier_breaks):
        end_roll = tier_breaks[i + 1][0] - 1
        roll_range = f"{start_roll}" if start_roll == end_roll else f"{start_roll}-{end_roll}"
    else:
        roll_range = f"{start_roll}+"
    entry = {"roll_range": roll_range}
    for j, inj in enumerate(injury_types):
        entry[inj.lower().replace(' ', '_').replace('/', '_').replace('&', 'and')] = int(vals[j]) if vals[j] is not None else None
    recovery.append(entry)

# ─────────────────────────────────────────────
# Build final JSON
# ─────────────────────────────────────────────

output = {
    "round": round_data,
    "actions": actions_out,
    "crit_size_adjustments": crit_size_out,
    "vision_penalties": vision_penalties,
    "slaying_bonuses": slaying_bonuses,
    "weapon_multi_attack_penalties": weapon_multi_attack,
    "mount_stats": mount_stats,
    "endurance_results": endurance_results,
    "recovery": recovery,
}

with open(OUTPUT, 'w', encoding='utf-8') as f:
    json.dump(output, f, indent=2, ensure_ascii=False)

# Summary
print("=== EXTRACTION SUMMARY ===")
print(f"round.ap_by_status: {len(ap_by_status)} entries")
print(f"round.notes: {len(notes)} notes")
print(f"actions categories: {len(actions_out)}")
for a in actions_out:
    print(f"  {a['category']}: {len(a['entries'])} entries")
print(f"crit_size_adjustments.size_table: {len(size_rows)} sizes")
print(f"crit_size_adjustments.crit_result_matrix: {len(crit_matrix)} crit levels")
print(f"vision_penalties: {len(vision_penalties)} conditions")
print(f"slaying_bonuses: {len(slaying_bonuses)} tiers")
print(f"weapon_multi_attack_penalties: {len(weapon_multi_attack)} entries")
print(f"mount_stats: {len(mount_stats)} mounts")
print(f"endurance_results: {len(endurance_results)} tiers")
print(f"recovery: {len(recovery)} roll tiers")
print(f"\nOutput written to: {OUTPUT}")
